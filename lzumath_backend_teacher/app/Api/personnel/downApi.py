from flask import Blueprint, request, jsonify, send_file
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Protection, Font, PatternFill, Border, Side
import io

templateDownload = Blueprint('templateDownload', __name__)

def connectDb():
    return pymysql.connect(
        host='student-mysql',
        user='root',
        password='123456',
        database='lzumathedu'
    )

@templateDownload.route('/api/template/download', methods=['POST'])
def template_download():
    data = request.get_json()
    college = data.get('college')
    grade = data.get('grade')
    major = data.get('major')
    tableName = f"studentsData_{college}_{grade}"

    try: 
        db = connectDb()
        cursor = db.cursor()
        sql = """SELECT table_name FROM information_schema.tables WHERE table_schema = %s AND table_name = %s"""
        cursor.execute(sql, ('lzumathedu', tableName))
        result = cursor.fetchone()

        if result:
            cursor.execute(f"SELECT * FROM `{tableName}` WHERE major = %s", (major,))
            rows = cursor.fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = f"{major}"
            headers = [
                "学号", "姓名", "学院", "年级", "专业", "综合成绩",
                "学业成绩", "学业成绩折合",
                "第二课堂", "学生互评", "素质测评折合",
                "附加分", "扣分"
            ]
            ws.append(headers)
            academic_ratio = 0.6

            first_data_row = 2
            for idx, row in enumerate(rows, start=first_data_row):
                sid = row[0]
                name = row[1]
                college = row[2]
                grade = row[3]
                major = row[4]
                
                ws.cell(row=idx, column=1, value=sid)     # 学号
                ws.cell(row=idx, column=2, value=name)    # 姓名
                ws.cell(row=idx, column=3, value=college) # 学院
                ws.cell(row=idx, column=4, value=grade)   # 年级
                ws.cell(row=idx, column=5, value=major)   # 专业
                
                # 综合成绩 = 学业成绩折合 + 素质评测折合 + 附加分 + 扣分
                ws.cell(row=idx, column=6, value=f"=H{idx}+K{idx}+L{idx}+M{idx}")

                ws.cell(row=idx, column=7, value="")      # 学业成绩
                ws.cell(row=idx, column=8, value=f"=G{idx}*$H${len(rows)+2}") # 学业成绩折合

                ws.cell(row=idx, column=9, value="")      # 第二课堂
                ws.cell(row=idx, column=10, value="")     # 学生互评
                ws.cell(row=idx, column=11, value=f"=(I{idx}/2+J{idx}/2)*$K${len(rows)+2}") # 素质测评折合

                ws.cell(row=idx, column=12, value=0)      # 附加分
                ws.cell(row=idx, column=13, value=0)      # 扣分

            ratio_row = len(rows) + 2
            ws.cell(row=ratio_row, column=7, value="学业成绩占比")
            ws.cell(row=ratio_row, column=8, value=academic_ratio)
            ws.merge_cells(start_row=ratio_row, start_column=9, end_row=ratio_row, end_column=10)
            ws.cell(row=ratio_row, column=9, value="素质评测占比")
            ws.cell(row=ratio_row, column=11, value=f"=1-H{ratio_row}")

            ws.protection.sheet = True # 启用工作表保护
            # 可编辑列：学业成绩(G=7)、第二课堂(I=9)、学生互评(J=10)、附加分(L=12)、扣分(M=13)
            editable_cols = [7, 9, 10, 12, 13]
            for r in range(2, len(rows) + 2):
                for c in editable_cols:
                    ws.cell(row=r, column=c).protection = Protection(locked=False)
            # 最后一行占比也可编辑
            ws.cell(row=ratio_row, column=8).protection = Protection(locked=False)

            # 可编辑区域高亮
            fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            for r in range(2, ratio_row + 1):
                for c in [7, 9, 10, 12, 13]:
                    cell = ws.cell(row=r, column=c)
                    if not cell.protection.locked:
                       cell.fill = fill
            ws.cell(row=ratio_row, column=8).fill = fill

            # 边框
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            for r in range(1, ratio_row):
                for c in range(1, 14):
                    ws.cell(row=r, column=c).border = border
            for r in range(ratio_row, ratio_row + 1):
                for c in range(7, 12):
                    ws.cell(row=r, column=c).border = border

            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            return send_file(
                file_stream,
                as_attachment=True,
                download_name=f"{college}_{grade}_{major}_综测模板.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        else:
            return jsonify({"message": "未检索到对应数据表"}), 404

    except Exception as e:
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500